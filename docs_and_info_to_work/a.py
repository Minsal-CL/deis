from openpyxl import load_workbook

# Ruta del archivo Excel
archivo_excel = "VocabularioNorma820.xlsx"

# Carga el archivo Excel
libro_excel = load_workbook(archivo_excel)

# Obtiene una lista de nombres de hojas
nombres_hojas = libro_excel.sheetnames

# Imprime la lista de nombres de hojas
print("Lista de hojas en el archivo Excel:")
for nombre_hoja in nombres_hojas:
    print(nombre_hoja)